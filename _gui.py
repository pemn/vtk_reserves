#!python
'''
Copyright 2017 - 2021 Vale

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

https://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.

*** You can contribute to the main repository at: ***

https://github.com/pemn/usage-gui
---------------------------------
'''

__version__ = 20211111

### { HOUSEKEEPING
import sys, os, os.path, time, logging
# fix for wrong path of pythoncomXX.dll in vulcan 10.1.5
if 'VULCAN_EXE' in os.environ:
  os.environ['PATH'] += ';' + os.environ['VULCAN_EXE'] + "/Lib/site-packages/pywin32_system32"
### } HOUSEKEEPING

### { UTIL
logging.basicConfig(format='%(message)s', level=99)
log = lambda *argv: logging.log(99, ' '.join(map(str,argv)))

def pyd_zip_extract():
  ''' embedded module unpacker '''
  zip_path = os.path.splitext(sys.argv[0])[0] + '.pyz'
  if not os.path.exists(zip_path):
    return
  platform_arch = '.cp%d%d-win_amd64' % (sys.hexversion >> 24, sys.hexversion >> 16 & 0xFF)

  pyd_path = os.environ.get('TEMP', '.') + "/pyz_" + platform_arch

  if not os.path.isdir(pyd_path):
    os.mkdir(pyd_path)

  sys.path.insert(0, pyd_path)
  os.environ['PATH'] = os.environ.get('PATH', '') + ';' + pyd_path

  import zipfile
  zip_root = zipfile.ZipFile(zip_path)

  for name in zip_root.namelist():
    m = re.match('[^/]+' + platform_arch + r'\.(pyd|zip)$', name, re.IGNORECASE)
    if m:
      if not os.path.exists(os.path.join(pyd_path, name)):
        zip_root.extract(name, pyd_path)
        if m.group(1) == "zip":
          subzip = os.path.join(pyd_path, name)
          zipfile.ZipFile(subzip).extractall(pyd_path)


def list_any(l):
  return sum(map(bool, l))

# subclass of list with a string representation compatible to perl argument input
# which expects a comma separated list with semicolumns between rows
class commalist(list):
  _rowfs = ";"
  _colfs = ","
  def parse(self, arg):
    "fill this instance with data from a string"
    if isinstance(arg, str):
      for row in arg.split(self._rowfs):
        self.append(row.split(self._colfs))
    else:
      self = commalist(arg)

    return self

  def __str__(self):
    r = ""
    # custom join
    for i in self:
      if(isinstance(i, list)):
        i = self._colfs.join(i)
      if len(r) > 0:
        r += self._rowfs
      r += i
    return r
  def __hash__(self):
    return len(self)

  # sometimes we have one element, but that element is ""
  # unless we override, this will evaluate as True
  def __bool__(self):
    return len(str(self)) > 0
  # compatibility if the user try to treat us as a real string
  def split(self, *args):
    return [",".join(_) for _ in self]

def pd_synonyms(df, synonyms, default = 0):
  ''' from a list of synonyms, find the best candidate amongst the dataframe columns '''
  if len(synonyms):
    # first try a direct match
    for v in synonyms:
      if v in df:
        return v
    # second try a case insensitive match
    for v in synonyms:
      m = df.columns.str.match(v, False)
      if m.any():
        return df.columns[m.argmax()]
  # fail safe to the first column
  if default is not None:
    return df.columns[default]
  return None

def table_name_selector(df_path, table_name = None):
  if table_name is None:
    m = re.match(r'^(.+)!(\w+)$', df_path)
    if m:
      df_path = m.group(1)
      table_name = m.group(2)

  return df_path, table_name

def bm_sanitize_condition(condition):
  if condition is None:
    condition = ""

  if len(condition) > 0:

    # convert a raw condition into a actual block select string
    if re.match(r'\s*\-', condition):
      # condition is already a select syntax
      pass
    elif re.search(r'\.00t$', condition, re.IGNORECASE):
      # bounding solid
      condition = '-X -t "%s"' % (condition)
    else:
      condition = '-C "%s"' % (condition.replace('"', "'"))

  return condition

# convert field names in the TABLE:FIELD to just FIELD or just TABLE
def table_field(args, table=False):
  if isinstance(args, list):
    args = [table_field(arg, table) for arg in args]
  elif args.find(':') != -1:
    if table:
      args = args[0:args.find(':')]
    else:
      args = args[args.find(':')+1:]
  return args

# wait and unlock block model
def bmf_wait_lock(path, unlock = False, tries = None):
    blk_lock = os.path.splitext(path)[0] + ".blk_lock"

    print("waiting lock", blk_lock)
    while os.path.isfile(blk_lock):
        if unlock and not tries:
          os.remove(blk_lock)
          print("removed lock", blk_lock)
          break
        
        if tries == 0:
          break
        if tries is not None:
          tries -= 1
          print("waiting lock", blk_lock, tries, "seconds")
          
        
        time.sleep(1)

### } UTIL

### { IO

def pd_load_dataframe(df_path, condition = '', table_name = None, vl = None, keep_null = False):
  '''
  convenience function to return a dataframe based on the input file extension
  csv: ascii tabular data
  xls: excel workbook
  bmf: vulcan block model
  dgd.isis: vulcan design object layer
  isis: vulcan generic database
  00t: vulcan triangulation
  dm: datamine generic database
  shp: ESRI shape file
  '''
  import pandas as pd
  if table_name is None:
    df_path, table_name = table_name_selector(df_path)
  df = None
  if not os.path.exists(df_path):
    print(df_path,"not found")
    df = pd.DataFrame()
  elif re.search(r'(csv|asc|prn|txt)$', df_path, re.IGNORECASE):
    df = pd.read_csv(df_path, None, engine='python', encoding="latin_1")
  elif re.search(r'xls\w?$', df_path, re.IGNORECASE):
    df = pd_load_excel(df_path, table_name)
  elif df_path.lower().endswith('bmf'):
    df = pd_load_bmf(df_path, condition, vl)
    condition = ''
  elif df_path.lower().endswith('dgd.isis'):
    df = pd_load_dgd(df_path, table_name)
  elif df_path.lower().endswith('isis'):
    df = pd_load_isisdb(df_path, table_name)
  elif df_path.lower().endswith('00t'):
    df = pd_load_tri(df_path)
  elif df_path.lower().endswith('00g'):
    df = pd_load_grid(df_path)
  elif df_path.lower().endswith('dm'):
    df = pd_load_dm(df_path)
  elif df_path.lower().endswith('shp'):
    df = pd_load_shape(df_path)
  elif df_path.lower().endswith('dxf'):
    df = pd_load_dxf(df_path)
  elif df_path.lower().endswith('json'):
    df = pd.read_json(df_path)
  elif df_path.lower().endswith('jsdb'):
    import jsdb_driver
    df = jsdb_driver.pd_load_database(df_path)
  elif df_path.lower().endswith('msh'):
    df = pd_load_mesh(df_path)
  elif df_path.lower().endswith('png'):
    df = pd_load_spectral(df_path)
  elif df_path.lower().endswith('obj'):
    df = pd_load_obj(df_path)
  elif re.search(r'tiff?$', df_path, re.IGNORECASE):
    import vulcan_save_tri
    df = vulcan_save_tri.pd_load_geotiff(df_path)
  else:
    df = pd.DataFrame()

  if not int(keep_null):
    df.mask(df == -99, inplace=True)

  # replace -99 with NaN, meaning they will not be included in the stats
  if len(condition):
    df.query(condition, True)

  return df

def pd_flat_columns(df):
  """ convert a column multi index into flat concatenated strings """
  fci = [" ".join(_) for _ in df.columns.to_flat_index()]
  df.set_axis(fci, 1, True)

def pd_save_dataframe(df, df_path, sheet_name='Sheet1'):
  import pandas as pd
  ''' save a dataframe to one of the supported formats '''
  if df.size:
    if df.ndim == 1:
      # we got a series somehow?
      df = df.to_frame()
    if not str(df.index.dtype).startswith('int'):
      levels = list(set(df.index.names).intersection(df.columns))
      if len(levels):
        df.reset_index(levels, drop=True, inplace=True)
      df.reset_index(inplace=True)
    if isinstance(df.columns, pd.MultiIndex):
      pd_flat_columns(df)
    if isinstance(df_path, pd.ExcelWriter) or df_path.lower().endswith('.xlsx'):
      # multiple excel sheets mode
      df.to_excel(df_path, index=False, sheet_name=sheet_name)
    elif df_path.lower().endswith('dgd.isis'):
      pd_save_dgd(df, df_path)
    elif df_path.lower().endswith('isis'):
      pd_save_isisdb(df, df_path)
    elif df_path.lower().endswith('bmf'):
      pd_save_bmf(df, df_path)
    elif df_path.lower().endswith('shp'):
      pd_save_shape(df, df_path)
    elif df_path.lower().endswith('dxf'):
      pd_save_dxf(df, df_path)
    elif df_path.lower().endswith('00t'):
      pd_save_tri(df, df_path)
    elif df_path.lower().endswith('json'):
      df.to_json(df_path, 'records')
    elif df_path.lower().endswith('jsdb'):
      import jsdb_driver
      jsdb_driver.pd_save_database(df, df_path)
    elif df_path.lower().endswith('msh'):
      pd_save_mesh(df, df_path)
    elif df_path.lower().endswith('obj'):
      pd_save_obj(df, df_path)
    elif df_path.lower().endswith('png'):
      pd_save_spectral(df, df_path)
    elif re.search(r'tiff?$', df_path, re.IGNORECASE):
      import vulcan_save_tri
      vulcan_save_tri.pd_save_geotiff(df, df_path)
    elif len(df_path):
      df.to_csv(df_path, index=False)
    else:
      print(df.to_string(index=False))
  else:
    print(df_path,"empty")

def dgd_list_layers(file_path):
  ''' return the list of layers stored in a dgd '''
  import vulcan
  r = []
  if vulcan.version_major < 11:
    db = vulcan.isisdb(file_path)
    r = [db.get_key() for _ in db.keys if db.get_key().find('$') == -1]
    db.close()
  else:
    dgd = vulcan.dgd(file_path)
    r = [_ for _ in dgd.list_layers() if _.find('$') == -1]
    dgd.close()

  return r

# Vulcan BMF
def bmf_field_list(file_path):
  import vulcan
  bm = vulcan.block_model(file_path)
  r = bm.field_list()
  bm.close()
  return r

def bm_get_pandas_proportional(self, vl=None, select=None):
  """
  custom get_pandas dropin replacement with proportional volume inside solid/surface
  """
  import pandas as pd
  if select is None:
     select = ''

  if vl is None:
     vl = self.field_list() + [ 'xlength', 'ylength', 'zlength', 'xcentre', 'ycentre', 'zcentre', 'xworld', 'yworld', 'zworld' ]
  
  vi = None
  if 'volume' in vl:
    vi = vl.index('volume')

  self.select(select)
  data = []
  for block in self:
    row = [self.get_string(v) if self.is_string( v ) else self.get(v) for v in vl]
    if vi is not None:
      row[vi] = self.match_volume()

    data.append(row)
  
  return pd.DataFrame(data, columns=vl)

def pd_load_bmf(df_path, condition = '', vl = None):
  import vulcan
  bm = vulcan.block_model(df_path)
  if vl is not None:
    vl = list(filter(bm.is_field, vl))
  
  # get a DataFrame with block model data
  if '-X' in condition:
    return bm_get_pandas_proportional(bm, vl, bm_sanitize_condition(condition))
  else:
    return bm.get_pandas(vl, bm_sanitize_condition(condition))

def pd_auto_schema(df, xyzd):
  import numpy as np
  import pandas as pd
  xyz0 = np.zeros(3)

  xyzo = np.min(df.loc[:, ['xworld','yworld','zworld']])
  xyz1 = np.subtract(np.max(df.loc[:, ['xworld','yworld','zworld']]), xyzo)
  xyz1 = np.multiply(np.ceil(np.divide(xyz1, xyzd)), xyzd)
  xyzo = np.subtract(xyzo, np.multiply(xyzd, [0.5,0.5,0.5]))
  xyzn = np.ceil(np.divide(xyz1, xyzd))
  return xyzo, xyz0, xyz1, xyzn

def pd_save_bmf(df, df_path):
  import numpy as np
  import pandas as pd
  xyzw = ['xworld','yworld','zworld']
  xyzd = np.ones(3)
  vlc = 3
  for v in xyzw:
    if v not in df:
      break
  else:
    if 'xlength' in df:
      vlc += 1
      xyzd[0] = df['xlength'].mean()
    if 'ylength' in df:
      vlc += 1
      xyzd[1] = df['ylength'].mean()
    if 'zlength' in df:
      vlc += 1
      xyzd[2] = df['zlength'].mean()

    xyzo, xyz0, xyz1, xyzn = pd_auto_schema(df, xyzd)

    vl = df.columns[vlc:]

    try:
      import vulcan
    except:
      print("vulcan module not found", file=sys.stderr)
      return
    bm = vulcan.block_model()
    # self, name, x0, y0, z0, x1, y1, z1, nx, ny, nz
    # bm.create_regular(df_path, *xyz0, *xyz1, *xyzn)
    # xyzn = xyzn.astype(np.int_)
    bm.create_regular(df_path, *xyz0, *xyz1, int(xyzn[0]), int(xyzn[1]), int(xyzn[2]))
    bm.set_model_origin(*xyzo)
    for v in vl:
      if df.dtypes[v] == 'object':
        bm.add_variable(v, 'name', 'n', '')
      else:
        bm.add_variable(v, 'float', '-99', '')
    bm.write()
    print("index model")
    bm.index_model()
    bm.write()

    for row in df.index:
      if row % 10000 == 0:
        print("blocks processed:",row)
      bm.find_world_xyz(*df.loc[row, xyzw])
      for v in vl:
        if df.dtypes[v] == 'object':
          bm.put_string(v, df.loc[row, v])
        else:
          bm.put(v, df.loc[row, v])
    print("blocks processed:",row)
    print(df_path, "saved")


# Vulcan ISIS database
def isisdb_list(file_path, alternate = False):
  import vulcan
  db = vulcan.isisdb(file_path)
  if alternate:
    vl = db.table_list()
  else:
    key = db.synonym('GEO','HOLEID')
    if not key:
      key = 'KEY'
    vl = db.field_list(db.table_list()[-1])
    vl.insert(0, key)
  db.close()
  return vl

def isisdb_check_table_name(db, table_name):
  # by default, use last table which is the desired one in most cases
  if table_name is None or table_name not in db.table_list():
    table_name = db.table_list()[-1]
  return table_name

def pd_load_isisdb(df_path, table_name = None):
  import pandas as pd
  if os.path.exists(df_path + '_lock'):
    raise Exception('Input database locked')
  import vulcan
  db = vulcan.isisdb(df_path)
  table_name = isisdb_check_table_name(db, table_name)

  vl = list(db.field_list(table_name))
  key = db.synonym('GEO','HOLEID')
  if not key:
    key = 'KEY'
  fdata = []
  db.rewind()
  while not db.eof():
    if table_name == db.get_table_name():
      row = [db[field] for field in vl]
      if key not in vl:
        row.insert(0, db.get_key())
      fdata.append(row)
    db.next()
  if key not in vl:
    vl.insert(0, key)
  return pd.DataFrame(fdata, None, vl)

def pd_save_isisdb(df, df_path, table_name = None):
  print("# pd_save_isisdb")
  import pandas as pd
  if os.path.exists(df_path + '_lock'):
    raise Exception('Input database locked')
  import vulcan
  db = vulcan.isisdb(df_path)
  table_name = isisdb_check_table_name(db, table_name)
  header = db.synonym('IDENT', 'HOLEID')
  if not header:
    header = db.table_list()[0]
  key = db.synonym('GEO','HOLEID')
  # default to first field of first table
  if not key:
    key = db.field_list(header)[0]
  if key not in df:
    df[key] = os.path.basename(df_path)
  vl = set(df.columns).intersection(db.field_list(table_name))
  print("header",header,"key",key,"table",table_name,"vl",vl)
  for gp,df_key in df.groupby(key):
    if db.find_key(gp) == 0:
      print("replaced",key,gp)
      db.delete_key(gp)
    else:
      print("created",key,gp)
    db.put_table_name(header)
    db.put_string(key, gp)
    db.append()
    db.put_table_name(table_name)
    for row in df_key.index:
      for v in vl:
        if db.is_string(v, table_name):
          db.put_string(v, str(df_key.loc[row, v]))
        else:
          db.put(v, df_key.loc[row, v])
      db.append()

def pd_update_isisdb(df, df_path, table_name = None, vl = None):
  import numpy as np
  import pandas as pd
  if os.path.exists(df_path + '_lock'):
    raise Exception('Input database locked')
  import vulcan
  if vl is None:
    vl = df.columns
  if np.ndim(vl) == 0:
    vl = [vl]

  db = vulcan.isisdb(df_path)
  table_name = isisdb_check_table_name(db, table_name)
  row = 0
  df.index = pd.RangeIndex(row,len(df))
  while not db.eof():
    if table_name == db.get_table_name():
      for v in vl:
        if db.is_string(v, table_name):
          db.put_string(v, str(df.loc[row, v]))
        else:
          db.put(v, float(df.loc[row, v]))
      row += 1
      if row >= len(df):
        break
    db.next()

def pd_load_dgd(df_path, layer_dgd = None):
  ''' create a dataframe with object points and attributes '''
  import pandas as pd
  import vulcan
  obj_attr = ['name', 'group', 'feature', 'description', 'value', 'colour']
  df = pd.DataFrame(None, columns=smartfilelist.default_columns + ['p','closed','layer','oid'] + obj_attr)

  dgd = vulcan.dgd(df_path)

  row = 0
  if dgd.is_open():
    layers = layer_dgd
    if layer_dgd is None:
      layers = dgd_list_layers(df_path)
    elif not isinstance(layer_dgd, list):
      layers = [layer_dgd]

    for l in layers:
      if not dgd.is_layer(l):
        continue
      layer = dgd.get_layer(l)
      oid = 0
      for obj in layer:
        df_row = pd.Series()
        df_row['oid'] = str(oid)
        df_row['layer'] = layer.get_name()
        for t in obj_attr:
          df_row[t] = getattr(obj, t)
        if obj.get_type() == 'TEXT3D':
          df_row['x'],df_row['y'],df_row['z'],df_row['w'],df_row['t'],df_row['p'] = obj.get_origin()
          df_row['n'] = 0
          df.loc[row] = df_row
          row += 1
        else:
          for n in range(obj.num_points()):
            df_row['closed'] = obj.is_closed()
            p = obj.get_point(n)
            df_row['x'] = p.get_x()
            df_row['y'] = p.get_y()
            df_row['z'] = p.get_z()
            df_row['w'] = p.get_w()
            df_row['t'] = p.get_t()
            df_row['p'] = p.get_name()
            # point sequence within this polygon
            df_row['n'] = n
            df.loc[row] = df_row
            row += 1
        oid += 1

  return df

def pd_save_dgd(df, df_path):
  ''' create vulcan objects from a dataframe '''
  import vulcan
  obj_attr = ['value', 'name', 'group', 'feature', 'description']
  xyz = ['x','y','z']
  if 'w' in df:
    xyz.append('w')
  if 't' in df:
    xyz.append('t')


  dgd = vulcan.dgd(df_path, 'w' if os.path.exists(df_path) else 'c')

  layer_cache = dict()

  c = []
  n = None
  for row in df.index[::-1]:
    layer_name = '0'
    if 'layer' in df:
      layer_name = df.loc[row, 'layer']
    if layer_name not in layer_cache:
      layer_cache[layer_name] = vulcan.layer(str(layer_name))

    if 'n' in df:    
      n = df.loc[row, 'n']
    else:
      n = row

    # last row special case
    c.insert(0, row)
    if n == 0:
      points = df.loc[c, xyz].values.tolist()
      obj = vulcan.polyline(points)
      if 'closed' in df:
        obj.set_closed(bool(df.loc[row, 'closed']))
      for i in range(len(obj_attr)):
        if obj_attr[i] in df:
          v = str(df.loc[row, obj_attr[i]])
          if i == 0:
            v = float(df.loc[row, obj_attr[i]])

          setattr(obj, obj_attr[i], v)

      layer_cache[layer_name].append(obj)
      c.clear()

  for v in layer_cache.values():
    dgd.save_layer(v)

# Vulcan Triangulation 00t
def pd_load_tri(df_path):
  import vulcan
  import numpy as np
  import pandas as pd
  tri = vulcan.triangulation(df_path)

  ta = vulcan.tri_attributes(df_path)
  

  cv = tri.get_colour()
  cn = 'colour'
  if vulcan.version_major >= 11 and tri.is_rgb():
    cv = np.sum(np.multiply(tri.get_rgb(), [2**16,2**8,1]))
    cn = 'rgb'
  #print("pd_load_tri nodes",tri.n_nodes(),"faces",tri.n_faces())
  #df = [tri.get_node(int(f[n])) + [0,bool(n),n,1,f[n],cv] for f in tri.get_faces() for n in range(3)], 
  # orphan nodes
  df =  nodes_faces_to_df(tri.get_vertices(), tri.get_faces())
  if ta.is_ok():
    for k,v in ta.get_hash().items():
      df[k] = v

  return df


def df_to_nodes_faces_simple(df, node_name = 'node', xyz = ['x','y','z']):
  ''' fast version only supporting a single triangulation '''
  import pandas as pd

  # nodes are duplicated for each face in df
  # create a list where they are unique again
  i_row = df[node_name].drop_duplicates().sort_values()
  nodes = df.loc[i_row.index, xyz]

  face_size = 3
  if 'n' in df and len(df):
    # handle orphan nodes
    face_n = df['n'].dropna()
    face_size = int(face_n.max()) + 1
    faces = df.loc[face_n.index, node_name].values.reshape((len(face_n) // face_size, face_size))
  else:
    # faces are unique so we can use the input data
    faces = df[node_name].values.reshape((len(df) // face_size, face_size))

  return nodes.values, faces

def df_to_nodes_faces_lines(df, node_name = 'node', xyz = ['x','y','z']):
  nodes = []
  lines = []
  if 'type' in df:
    l_nodes, lines = df_to_nodes_lines(df.query("type != 'TRIANGLE'"), node_name, xyz)
    nodes.extend(l_nodes)
    df = df.query("type == 'TRIANGLE'")

  t_nodes, faces = df_to_nodes_faces(df, node_name, xyz)

  nodes.extend(t_nodes)
  return nodes, faces, lines

def df_to_nodes_lines(df, node_name, xyz):
  nodes = df[xyz].values.take(df[node_name], 0)
  lines = []
  n = len(df)
  part = []
  for row in df.index[::-1]:
    if 'n' in df:
      n = df.loc[row, 'n']
    else:
      n -= 1
    part.insert(0, n)
    if n == 0:
      lines.insert(0, part)
      part = []
  return nodes, lines

df_to_nodes_faces = df_to_nodes_faces_simple

def pd_save_tri(df, df_path):
  import vulcan
  import numpy as np
  
  if os.path.exists(df_path):
    os.remove(df_path)

  tri = vulcan.triangulation("", "w")

  if 'rgb' in df:
    rgb = np.floor(np.divide(np.mod(np.repeat(df.loc[0, 'rgb'],3), [2**32, 2**16, 2**8]), [2**16,2**8,1]))
    print('color r %d g %d b %d' % tuple(rgb))
    tri.set_rgb(rgb.tolist())
  elif 'colour' in df:
    print('colour index ', df.loc[0, 'colour'])
    tri.set_colour(int(df.loc[0, 'colour']))
  else:
    print('default color')
    tri.set_colour(1)

  if 'filename' not in df:
    df['filename'] = ''

  nodes, faces = df_to_nodes_faces(df)
  print("nodes",len(nodes))
  print("faces",len(faces))
  for n in nodes:
    tri.add_node(*n)
  for f in faces:
    tri.add_face(*map(int,f))

  tri.save(df_path)

# Vulcan Grid 00g
def pd_load_grid(df_path):
  import vulcan

  grid = vulcan.grid(df_path)
  df = grid.get_pandas()
  df['filename'] = os.path.basename(df_path)
  return df

# Datamine DM
def pd_load_dm(df_path, condition = ''):
  import win32com.client
  import pandas as pd
  dm = win32com.client.Dispatch('DmFile.DmTable')
  dm.Open(df_path, 0)
  fdata = []
  n = dm.Schema.FieldCount
  for i in range(dm.GetRowCount()):
    #fdata.append([dm.GetColumn(j) for j in range(1, n)])
    row = [None] * n
    for j in range(n):
      v = dm.GetColumn(j+1)
      if v != dm.Schema.SpecialValueAbsent:
        row[j] = v
    fdata.append(row)
    dm.GetNextRow()
  
  return pd.DataFrame(fdata, None, [dm.Schema.GetFieldName(j+1) for j in range(n)])

def dm_field_list(file_path):
  import win32com.client
  dm = win32com.client.Dispatch('DmFile.DmTable')
  dm.Open(file_path, 0)
  r = [dm.Schema.GetFieldName(j) for j in range(1, dm.Schema.FieldCount + 1)]
  return r

# Microsoft Excel compatibles
def excel_field_list(df_path, table_name, alternate = False):
  r = []
  try:
    import openpyxl
    wb = openpyxl.load_workbook(df_path)
    if alternate:
      r = wb.sheetnames
    elif table_name and table_name in wb:
      r = next(wb[table_name].values)
    else:
      r = next(wb.active.values)
  except:
    print("openpyxl not available")
    import pandas as pd
    r = pd.read_excel(df_path).columns
  
  return r

def pd_load_excel_350(df_path, table_name):
  import pandas as pd
  import openpyxl
  wb = openpyxl.load_workbook(df_path)
  if table_name and table_name in wb:
    data = wb[table_name].values
  else:
    data = wb.active.values
  cols = next(data)
  return pd.DataFrame(data, columns=[i if cols[i] is None else cols[i] for i in range(len(cols))])

def pd_load_excel(df_path, table_name = None):
  if sys.hexversion < 0x3060000:
    return pd_load_excel_350(df_path, table_name)
  import pandas as pd
  df = None
  engine='openpyxl'
  if df_path.lower().endswith('.xls'):
    # openpyxl is not handling xls anymore
    engine = None
  df = pd.read_excel(df_path, table_name, engine=engine)
  if not isinstance(df, pd.DataFrame):
    _, df = df.popitem()

  return df

def pd_save_excel_tables(save_path, *arg):
  import openpyxl
  from openpyxl.worksheet.table import Table
  from openpyxl.utils import get_column_letter
  from openpyxl.utils.dataframe import dataframe_to_rows

  wb = openpyxl.Workbook()
  # remove the default Sheet1
  wb.remove(wb.active)
  for i in range(0, len(arg), 2):
    t = None
    if i+1 < len(arg):
      t = arg[i+1]
    if t is None:
      t = 'Table%d' % (i / 2 + 1)
    ws = wb.create_sheet(title=t)
    df = arg[i]
    if not str(arg[i].index.dtype).startswith('int'):
      df = arg[i].reset_index()
    for r in dataframe_to_rows(df, index=False, header=True):
      ws.append(r)

    tab = Table(displayName=t, ref="A1:%s%d" % (get_column_letter(ws.max_column), ws.max_row))
    ws.add_table(tab)

  wb.save(save_path)

# ESRI shape
def pd_load_shape(file_path):
  import pandas as pd
  import shapefile

  shapes = shapefile.Reader(file_path, encodingErrors='replace')
  columns = smartfilelist.default_columns + ['oid','part','type','layer']
  record_n = 0
  row = 0
  rows = []
  for item in shapes.shapeRecords():
    # object without a valid layer name will have this default layer
    fields = item.record.as_dict()

    p1 = len(item.shape.points)
    # each object may have multiple parts
    # create a object for each of these parts
    part_n = len(item.shape.parts)
    parts = item.shape.parts
    has_z = hasattr(item.shape, 'z')
    # handle the case where there are no parts (point primitives, usually)
    if len(parts) == 0:
      parts = [0]
    print(record_n, item.shape.shapeTypeName, p1)
    r = pd.Series()
    r['oid'] = record_n
    r['type'] = item.shape.shapeTypeName
    r['w'] = 0
    points = item.shape.points
    for k,v in fields.items():
      r[k] = v
    for p in reversed(parts):
      part_n -= 1
      for n in range(p,p1):
        if has_z:
          r['z'] = item.shape.z[n]
        # x,y
        for c in range(len(points[n])):
          r[columns[c]] = points[n][c]
        r['n'] = n - p
        r['t'] = n != p
        rows.append(r.copy())
        row += 1
      p1 = p
    record_n += 1

  return pd.DataFrame.from_records(rows)

def pd_save_shape(df, df_path):
  import shapefile
  import numpy as np
  import pandas as pd
  shpw = shapefile.Writer(os.path.splitext(df_path)[0], encodingErrors='replace')

  rc = []
  for i in range(df.shape[1]):
    if df.columns[i] not in 'xyzwtn':
      shpw.field(df.columns[i], 'C' if df.dtypes[i] == 'object' else 'F', decimal=4)
      rc.append(df.columns[i])

  p = []
  n = len(df)
  xyzwt = [_ for _ in 'xyz' if _ in df]
  for row in df.index[::-1]:
    if 'n' in df:
      n = df.loc[row, 'n']
    else:
      n -= 1

    p.insert(0, row)
    if n == 0:
      pdata = df.loc[p, xyzwt].values.tolist()
      ptype = ''
      if 'type' in df:
        ptype = df['type'].max()
      #if ptype == 'LINEARRING':
      #  shpw.polyz([pdata])
      if ptype.find('LINE') >= 0:
        shpw.linez([pdata])
      elif ptype.find('POINT') >= 0:
        if len(pdata) == 1:
          shpw.pointz(pdata[0][0], pdata[0][1], pdata[0][2])
        else:
          shpw.multipointz(pdata)
      else:
        shpw.polyz([pdata])

      shpw.record(*[np.nan_to_num(df.loc[row, c]) for c in rc])
      p.clear()
  
  shpw.close()

def shape_field_list(file_path):
  import shapefile
  shapes = shapefile.Reader(file_path)
  return [shapes.fields[i][0] for i in range(1, len(shapes.fields))]

# DXF
def pd_load_dxf(df_path):
  import pandas as pd
  import ezdxf
  r = []
  doc = ezdxf.readfile(df_path)
  # iterate over all entities in modelspace
  msp = doc.modelspace()
  for e in msp:
    print(e.dxf.layer, e.dxftype(), e.dxf.handle)
    n = 0
    mode = ''
    text = ''
    is_closed = 1
    points = []
    if e.dxftype() == 'POLYLINE':
      points = e.points()
      mode = e.get_mode()
      is_closed = e.is_closed
    elif e.dxftype() == 'TEXT':
      points.append(e.insert)
      text = e.plain_text()
    elif e.dxftype() == 'POINT':
      points.append(e.dxf.location)
      is_closed = 0
    elif hasattr(e.dxf, 'elevation'):
      z = e.dxf.elevation
      if isinstance(z, tuple):
        z = sum(z)
      points = [(_[0], _[1], z) for _ in e.get_points()]
      is_closed = e.is_closed
    else:
      points = e.get_points()
      is_closed = e.is_closed

    for p in points:
      r.append(tuple(p) + (n, is_closed, e.dxf.layer, e.dxftype(), mode, text, e.dxf.handle, e.dxf.color))
      n += 1
  c = ('x','y','z','n','closed','layer', 'type', 'mode','entityhandle', 'text', 'color')
  return pd.DataFrame(r, columns=c)

def pd_save_dxf(df, df_path):
  import pandas as pd
  import ezdxf
  xyz = ['x','y']
  if 'z' in df:
    xyz.append('z')

  doc = ezdxf.new(setup=True)
  msp = doc.modelspace()

  n = len(df)
  p = []
  for row in df.index[::-1]:
    if 'n' in df:
      n = df.loc[row, 'n']
    else:
      n -= 1

    p.insert(0, row)

    if n == 0:
      attribs = dict([(_, df.loc[row, _]) for _ in ['layer','color','closed'] if _ in df])
      pdata = df.loc[p, xyz].values.tolist()
      ptype = ''
      msp.add_polyline3d(pdata, dxfattribs=attribs)
      p.clear()

  doc.saveas(df_path)

# Leapfrog Mesh
def leapfrog_load_mesh(df_path):
  import struct
  file = open(df_path, "rb")

  index = None
  binary = None
  # read the structures from index section
  while True:
    if binary is None:
      line = file.readline()
      if line.startswith(b'[index]'):
        index = []
      elif line.startswith(b'[binary]'):
        binary = line[8:]
      elif index is not None:
        m = re.findall(rb'(\w+) (\w+) (\d+) ?(\d*)', line)
        if m:
          index.append(m[0])
    else:
      line = file.read(0xffff)
      binary += line
      if len(line) < 0xffff:
        break

    if len(line) == 0:
      break
  file.close()

  log("%02x%02x%02x%02x %02x%02x%02x%02x %02x%02x%02x%02x = %.2f %.2f %.2f" % tuple(struct.unpack_from('12B', binary, 0) + struct.unpack_from('3f', binary, 0)))
  # skip unknown 12 byte header
  # maybe on some cases it contains rgb color?
  p = 12
  store = {}
  # extract from the binary section the data for each structure 
  for part_name, part_type, part_wide, part_size in index:
    # when part_wide is 1 its ommited in the msh header
    if not part_size:
      part_size = part_wide
      part_wide = 1
    part_name = str(part_name, encoding='ascii')
    part_type = str(part_type, encoding='ascii').lower()
    part_wide = int(part_wide)
    part_size = int(part_size)

    part_pack = struct.Struct(part_type[0] * part_wide)
    part_data = []
    for i in range(part_size):
      part_data.append(part_pack.unpack_from(binary, p))
      p += part_pack.size
    store[part_name] = part_data

  return store.get('Location', []), store.get('Tri', [])

def nodes_faces_to_df(nodes, faces):
  import pandas as pd
  uniquenodes = set()
  for f in faces:
    uniquenodes.update(f)

  df_data = [tuple(nodes[int(f[n])]) + (0,bool(n),n,1,f[n]) for f in faces for n in range(3)]
  # handle orphan points
  if len(nodes) > len(uniquenodes):
    orphannodes = uniquenodes.symmetric_difference(range(len(nodes)))
    print("orphan nodes:", len(orphannodes))
    for i in orphannodes:
      df_data.append(tuple(nodes[int(i)]) + (0,0,None,1,i))

  return pd.DataFrame(df_data, columns=smartfilelist.default_columns + ['closed','node'])

def pd_load_mesh(df_path):
  nodes, faces = leapfrog_load_mesh(df_path)
  print("nodes",len(nodes))
  print("faces",len(faces))
  return nodes_faces_to_df(nodes, faces)
  
def leapfrog_save_mesh(nodes, faces, df_path):
  import struct
  import numpy as np

  file = open(df_path, "wb")
  file.write(b'%%ARANZ-1.0\n\n[index]\nTri Integer 3 %d;\nLocation Double 3 %d;\n\n[binary]' % (len(faces), len(nodes)))
  # write unknown header
  file.write(struct.pack('3i', 15732735, 1115938331, 1072939210))

  file.write(struct.pack(str(len(faces) * 3) + 'i', *np.ravel(faces)))
  file.write(struct.pack(str(len(nodes) * 3) + 'd', *np.ravel(nodes)))

  file.close()

def pd_save_mesh(df, df_path):
  nodes, faces = df_to_nodes_faces(df)
  leapfrog_save_mesh(nodes, faces, df_path)

def img_to_df(img):
  import numpy as np
  import pandas as pd
  channels = 1
  if img.ndim >= 3:
    channels = img.shape[2]
  dfi = np.indices(img.shape[:2]).transpose(1,2,0).reshape((np.prod(img.shape[:2]),2))
  dfx = img.reshape((np.prod(img.shape[:2]), channels))
  return pd.DataFrame(np.concatenate([dfi, dfx], 1), columns=['x','y'] + list(map(str,range(channels))))

# images and other binary databases
def pd_load_spectral(df_path):
  import skimage.io
  return img_to_df(skimage.io.imread(df_path))

def pd_save_spectral(df, df_path):
  import skimage.io
  import numpy as np
  # original image width and height are recoverable from the max x and max y
  wh = np.max(df, 0)
  df.drop(['x','y'], 1, inplace=True)

  if wh.size > 3:
    im_out = np.reshape(df.values, (wh['x'] + 1, wh['y'] + 1, wh.size - 2))
  else:
    im_out = np.reshape(df[df.columns[0]], (wh['x'] + 1, wh['y'] + 1))
  skimage.io.imsave(df_path, im_out)


### Wavefront OBJ ###
def wavefront_load_obj(df_path):
  od = {"v": [], "f": [], "l": [], "vt": []}
  file = open(df_path, "r")
  for l in file:
    c = l.split()
    if c[0] == 'vt':
      od['vt'].append(tuple(map(float, c[1:])))
    elif c[0] == 'v':
      od['v'].append(tuple(map(float, c[1:])))
    if c[0] == 'f':
      od['f'].append([int(_) - 1 for _ in c[1:]])
    if c[0] == 'l':
      od['l'].append([int(_) - 1 for _ in c[1:]])
  file.close()
  return od

def pd_load_obj(df_path):
  import pandas as pd
  od = wavefront_load_obj(df_path)
  nodes = od.get('v')
  faces = od.get('f')
  lines = od.get('l')
  df = pd.DataFrame()
  columns = smartfilelist.default_columns + ['closed','node','type']
  if len(faces):
    # mesh
    df = df.append(pd.DataFrame([nodes[int(f[i])] + (0,bool(i),i,1,int(f[i]),'TRIANGLE') for f in faces for i in range(len(f))], columns=columns))
  if len(lines):
    # polyline
    df = df.append(pd.DataFrame([nodes[int(n)] + (0,bool(n),n,1,int(n),'POLYLINE') for f in lines for n in f], columns=columns))
  if len(lines) == 0 and len(faces) == 0:
    # point cloud
    df = df.append(pd.DataFrame([p + (0,0,0,0,0,'POINT') for p in nodes], columns=columns))

  return df

def wavefront_save_obj(df_path, od):
  ' serialize a wavefront mesh dict to a file '
  file = open(df_path, "w")
  for k,v in od.items():
    for n in v:
      if k in ['f','l']:
        print(k, *[_ + 1 for _ in n], file=file)
      else:
        print(k, *n, file=file)
  file.close()

def pd_save_obj(df, df_path):
  nodes, faces, lines = df_to_nodes_faces_lines(df)
  wavefront_save_obj(df_path, {"v": nodes, "f": faces, "l": lines})

### } IO

### { GUI
import tkinter as tk
import tkinter.ttk as ttk
import tkinter.messagebox as messagebox
import tkinter.filedialog as filedialog
import re, pickle, threading
from PIL import Image, ImageDraw

# show gui or run main
def usage_gui(usage = None):
  '''
  this function handles most of the details required when using a exe interface
  '''
  # we already have arguments, run business code
  if(len(sys.argv) > 1):
    # traps help switches: /? -? /h -h /help -help
    if(usage is not None and re.match(r'[\-/](?:\?|h|help)$', sys.argv[1])):
      print(usage)
    else:
      # this special main will redirect to the business script
      main(*sys.argv[1:])

  else:
    AppTk(usage).mainloop()

class ClientScript(list):
  '''Handles the script with the same name as this interface file'''
  # magic signature that a script file must have for defining its gui
  _magic = r"usage:\s*\S+\s*([^\"\'\\]+)"
  _usage = None
  _type = None
  _file = None
  _base = None
  @classmethod
  def init(cls, client):
    cls._file = client
    cls._base = os.path.splitext(cls._file)[0]
    # HARDCODED list of supporte file types
    # to add a new file type, just add it to the list
    for ext in ['csh','lava','pl','bat','vbs','js']:
      if os.path.exists(cls._base + '.' + ext):
        cls._file = cls._base + '.' + ext
        cls._type = ext.lower()
        break

  @classmethod
  def exe(cls):
    if cls._type == "csh":
      return ["csh","-f"]
    if cls._type == "bat":
      return ["cmd", "/c"]
    if cls._type == "vbs" or cls._type == "js":
      return ["cscript", "/nologo"]
    if cls._type == "lava" or cls._type == "pl":
      return ["perl"]
    if cls._type is None:
      return ["python"]
    return []

  @classmethod
  def run(cls, script):
    print("# %s %s started" % (time.strftime('%H:%M:%S'), cls.file()))
    p = None
    if cls._type is None:
      import multiprocessing
      p = multiprocessing.Process(None, main, None, script.get())
      p.start()
      p.join()
      p = p.exitcode
    else:
      import subprocess
      # create a new process and passes the arguments on the command line
      args = cls.exe() + [cls._file] + script.getArgs()
      p = subprocess.Popen(" ".join(args))
      p.wait()
      p = p.returncode

    if not p:
      print("# %s %s finished" % (time.strftime('%H:%M:%S'), cls.file()))
    return p

  @classmethod
  def type(cls):
    return cls._type
  
  @classmethod
  def base(cls):
    return cls._base
  
  @classmethod
  def file(cls, ext = None):
    if ext is not None:
      return cls._base + '.' + ext
    return os.path.basename(cls._file)
  
  @classmethod
  def args(cls, usage = None):
    r = []
    if usage is None and cls._type is not None:
      usage = cls.parse()

    if usage:
      m = re.search(cls._magic, usage, re.IGNORECASE)
      if(m):
        cls._usage = m.group(1)
    
    if cls._usage is None or len(cls._usage) == 0:
      r = ['arguments']
    else:
      r = cls._usage.split()
    return r

  @classmethod
  def fields(cls, usage = None):
    return [re.match(r"^\w+", _).group(0) for _ in cls.args(usage)]

  @classmethod
  def parse(cls):
    if os.path.exists(cls._file):
      with open(cls._file, 'r', encoding='latin_1') as file:
        for line in file:
          if re.search(cls._magic, line, re.IGNORECASE):
            return line
    return None

  @classmethod
  def header(cls):
    r = ""
    if os.path.exists(cls._file):
      with open(cls._file, 'r') as file:
        for line in file:
          if(line.startswith('#!')):
            continue
          m = re.match(r'#\s*(.+)', line)
          if m:
            r += m.group(1) + "\n"
          else:
            break
    return r

class Settings(str):
  '''provide persistence for control values using pickled ini files'''
  _ext = '.ini'
  def __new__(cls, value=''):
    if len(value) == 0:
      value = os.path.splitext(os.path.realpath(sys.argv[0]))[0]
    if not value.endswith(cls._ext):
      value += cls._ext

    return super().__new__(cls, value)

  def save(self, obj):
    pickle.dump(obj, open(self,'wb'), 4)
    
  def load(self):
    if os.path.exists(self):
      return pickle.load(open(self, 'rb'))
    return {}

# option lists
class smartfilelist(object):
  '''
  detects file type and return a list of relevant options
  searches are cached, so subsequent searcher for the same file path are instant
  '''
  default_columns = ['x','y','z','w','t','n']
  # global value cache, using path as key
  _cache = [{},{}]
  @staticmethod
  def get(df_path, s = 0):
    import pandas as pd
    # special case for multiple files. use first
    if isinstance(df_path, commalist):
      if len(df_path):
        df_path = df_path[0][0]
      else:
        df_path = ""
    
    r = []
    # if this file is already cached, skip to the end
    if(df_path in smartfilelist._cache[s]):
      r = smartfilelist._cache[s][df_path]
    else:
      df_path, table_name = table_name_selector(df_path)
      if os.path.exists(df_path):
        input_ext = os.path.splitext(df_path.lower())[1]
        if df_path.lower().endswith(".dgd.isis"):
          if s == 1:
            r = dgd_list_layers(df_path)
          else:
            r = smartfilelist.default_columns + ['p','closed','layer','oid','name','group','feature','description','value','colour']
        elif input_ext == ".isis":
          r = isisdb_list(df_path, s)
        elif input_ext == ".bmf":
          r = bmf_field_list(df_path)
        elif input_ext == ".00t" and s == 0:
          r = smartfilelist.default_columns + ['closed','node','rgb','colour']
        elif input_ext == ".00g" and s == 0:
          r = ['x','y','value','mask','filename']
        elif input_ext == ".msh" and s == 0:
          r = smartfilelist.default_columns + ['closed','node']
        elif input_ext == ".csv":
          df = pd.read_csv(df_path, encoding="latin_1", nrows=s == 0 and 1 or None)
          if s == 0:
            r = df.columns.tolist()
          if s == 1:
            r = df[df.columns[0]].tolist()
        elif input_ext == ".json":
          df = pd.read_json(df_path)
          if s == 1:
            r = df['name'].tolist()
          else:
            r = df.columns.tolist()
        elif input_ext == ".ipynb":
          import json
          d = json.load(open(df_path, 'rb'))
          i = 0
          while i < len(d['cells']):
            if d['cells'][i]['cell_type'] == 'code':
              break
          r = [str.split(_, ' = ')[0] for _ in  d['cells'][i]['source']]
        elif re.search(r'xls\w?$', df_path, re.IGNORECASE):
          r = excel_field_list(df_path, table_name, s)
        elif input_ext == ".dm" and s == 0:
          r = dm_field_list(df_path)
        elif input_ext == ".shp" and s == 0:
          r = shape_field_list(df_path)
        elif input_ext == ".obj" and s == 0:
          r = smartfilelist.default_columns + ['closed','node']
        elif input_ext == ".dxf" and s == 0:
          r = smartfilelist.default_columns + ['layer']
        elif input_ext == ".zip" and s == 0:
          from zipfile import ZipFile
          r = ZipFile(df_path).namelist()
        elif input_ext == ".png" and s == 0:
          r = list('xy0123456789')
        elif re.search(r'tiff?$', df_path, re.IGNORECASE):
          r = ['x', 'y', 'x0', 'y0', '0', '1', '2', '3']
        elif input_ext == ".vtk":
          r.append('volume')
          r.append('region')
          try:
            import pyvista as pv
            r.extend(pv.read(df_path).array_names)
          except:
            print("pyvista or vtk modules could not be loaded")

        smartfilelist._cache[s][df_path] = r

    return r

# helper class to store parsed control definition tokens
class UsageToken(str):
  '''handles the token format used to creating controls'''
  _name = None
  _type = None
  _data = None
  def __init__(self, arg):
    super().__init__()
    # parse the token string extracting:
    # 1 name of the control
    # 2 type of the control which is a single character symbol
    # 3 any arguments which may be required for that token
    # the arguments may be another control name, default values 
    # or even specific constants which are expected by some controls
    m = re.match(r"(\w*)(\*|@|#|=|:|%|~|!|\?)(.*)", arg)
    if (m):
      self._name = m.group(1)
      self._type = m.group(2)
      self._data = m.group(3)
    else:
      self._name = arg
  @property
  def name(self):
    return self._name
  @property
  def type(self):
    return self._type
  @property
  def data(self):
    return self._data

# main content of dynamic form
class ScriptFrame(ttk.Frame):
  '''frame that holds the script argument controls'''
  _tokens = None
  def __init__(self, master, usage = None):
    ttk.Frame.__init__(self, master)

    self._tokens = [UsageToken(_) for _ in ClientScript.args(usage)]
    # for each token, create a child control of the apropriated type
    for token in self._tokens:
      c = None
      if token.type == '@':
        c = CheckBox(self, token.name, int(token.data) if token.data else 0)
      elif token.type == '*':
        c = FileEntry(self, token.name, token.data)
      elif token.type == '=':
        c = LabelCombo(self, token.name, token.data)
      elif token.type == '#':
        c = tkTable(self, token.name, token.data.split('#'))
      elif token.type == '%':
        c = LabelRadio(self, token.name, token.data)
      elif token.type == '!':
        c = ComboPicker(self, token.name, token.data, True)
      elif token.type == '?':
        if len(token.data):
          c = HiddenInput(self, token.name, token.data)
        else:
          c = CredentialsInput(self, token.name)
      elif token.type == ':':
        if token.data == 'portal':
          import gisportal
          c = gisportal.ArcGisField(self, token.name, token.data)
        else:
          c = ComboPicker(self, token.name, token.data)
      elif token.type == '~':
        import gisportal
        c = gisportal.ArcGisPortal(self, token.name, None, token.data)
      elif token.name:
        c = LabelEntry(self, token.name)
      else:
        continue
      c.pack(anchor="w", padx=20, pady=10, fill=tk.BOTH)
      
  def copy(self):
    "Assemble the current parameters and copy the full command line to the clipboard"
    cmd = " ".join(ClientScript.exe() + [ClientScript.file()] + self.getArgs())
    print(cmd)
    self.master.clipboard_clear()
    self.master.clipboard_append(cmd)
    # workaround due to tkinter clearing clipboard on exit
    messagebox.showinfo(message='Command line copied to clipboard.\nWill be cleared after interface closes.')
  
  @property
  def tokens(self):
    return self._tokens

  # get panel parameters as a list of strings
  def get(self, labels=False):
    if labels:
      return dict([[k, v.get()] for k,v in self.children.items()])
    return [self.children[t.name].get() for t in self.tokens]
  
  # get panel parameters as a flat string
  def getArgs(self):
    args = []
    for t in self.tokens:
      arg = str(self.children[t.name].get())
      if len(arg) == 0 or not set(' ";%\\').isdisjoint(arg):
        arg = '"' + arg + '"'
      args.append(arg)

    return args
  
  def set(self, values):
    if isinstance(values, dict):
      for k,v in self.children.items():
        if k in values and str(values[k]) != 'nan':
          v.set(values[k])
    else:
      for i in range(len(self.tokens)):
        self.children[self.tokens[i].name].set(values[i])

class HiddenInput(ttk.Frame):
  ''' a hidden frame hosting a value defined in usage string '''
  def __init__(self, master, label, source):
    ttk.Frame.__init__(self, master, name=label)
    self._variable = tk.StringVar(value=source)

  def get(self):
    return self._variable.get()
    
  def set(self, value):
    return None

class LabelEntry(ttk.Frame):
  ''' should behave the same as Tix LabelEntry but with some customizations '''
  _label = None
  _control = None
  def __init__(self, master, label):
    # create a container frame for the combo and label
    ttk.Frame.__init__(self, master, name=label)
    
    if isinstance(master, tkTable):
      self._control = ttk.Entry(self)
    else:
      self._control = ttk.Entry(self)
      self._label = ttk.Label(self, text=label, width=-20)
      self._label.pack(side=tk.LEFT)

    self._control.pack(expand=True, fill=tk.BOTH, side=tk.RIGHT)
    
  def get(self):
    return self._control.get()
   
  def set(self, value):
    if(value == None or len(value) == 0):
      return
    self._control.delete(0, tk.END)
    self._control.insert(0, value)

  def configure(self, **kw):
    if self._label is not None:
      self._label.configure(**kw)
    self._control.configure(**kw)
  
class LabelRadio(ttk.Labelframe):
  def __init__(self, master, label, source):
    self._variable = tk.StringVar()
    ttk.Labelframe.__init__(self, master, name=label, text=label)
    for _ in source.split(','):
      ttk.Radiobutton(self, variable=self._variable, text=_, value=_).pack(anchor="w")

  def get(self):
    return self._variable.get()
    
  def set(self, value):
    return self._variable.set(value)

  def configure(self, **kw):
    if "state" in kw:
      for v in self.children.values():
        v.configure(**kw)
    else:
      super().configure(**kw)

# checkbox + Var
class CheckBox(ttk.Checkbutton):
  '''superset of checkbutton with a builtin variable'''
  def __init__(self, master, label, reach=0):
    self._variable = tk.BooleanVar()
    self._reach = reach
    ttk.Checkbutton.__init__(self, master, name=label, text=label, variable=self._variable)
    self.bind("<ButtonPress>", self.onButtonPress)
    self.bind("<Configure>", self.onButtonPress)

  def onButtonPress(self, event=None):
    if self._reach > 0:
      value = self.get()
      # invert the selection when caller is the onclick
      # because the current value is the oposite of the future value
      if int(event.type) == 4:
        value = not value
      bubble = 0
      for v in sorted(self.master.children.values(), key=tk.Misc.winfo_y):
        if(v is self):
          bubble = self._reach
        elif(bubble > 0):
          bubble -= 1
          v.configure(state = "enabled" if value else "disabled")

  def get(self):
    return int(self._variable.get())
    
  def set(self, value):
    return self._variable.set(value)

# shorten paths when they are subdirectories of the current working dir
def relative_paths(path):
    cwd_drive, cwd_tail = os.path.splitdrive(os.getcwd().lower())
    path_drive, path_tail = os.path.splitdrive(path.lower())
    if cwd_drive == path_drive and os.path.commonpath([path_tail, cwd_tail]) == cwd_tail:
      return os.path.relpath(path)
    return(path)

# label + entry + dropdown list
class LabelCombo(ttk.Frame):
  _label = None
  _control = None
  def __init__(self, master, label, source=None):
    ttk.Frame.__init__(self, master, name=label)
    self._source = source
    if isinstance(master, tkTable):
      self._control = ttk.Combobox(self)
    else:
      self._control = ttk.Combobox(self, width=-60)
      self._label = ttk.Label(self, text=label, width=-20)
      self._label.pack(side=tk.LEFT)

    self._control.pack(expand=True, fill=tk.BOTH, side=tk.RIGHT)
    if source is not None:
      self.setValues(source.split(","))

  def get(self):
    return self._control.get()
  
  def set(self, value):
    return self._control.set(value)
  
  def setValues(self, values):
    self._control['values'] = values
    # MAGIC: if only one value in the list, use it as default
    if (len(values) == 1):
      self.set(values[0])
    # MAGIC: if any of the values is the same name as the control, select it
    for _ in values:
      if _.lower() == self.winfo_name():
        self.set(_)

  def configure(self, **kw):
    if self._label is not None:
      self._label.configure(**kw)
    self._control.configure(**kw)

class ComboPicker(LabelCombo):
  def __init__(self, master, label, source, alternate = False):
    LabelCombo.__init__(self, master, label)
    self._source = source
    self._alternate = alternate
    self._control.bind("<ButtonPress>", self.onButtonPress)
    
  def onButtonPress(self, *args):
    # temporarily set the cursor to a hourglass
    self._control['cursor'] = 'watch'
    source_widget = None
    if (isinstance(self.master, tkTable)):
      if (self._source in self.master.master.children):
        source_widget = self.master.master.nametowidget(self._source)
      else:
        _,_,row = self.winfo_name().rpartition("_")
        if self._source + "_" + row in self.master.children:
          source_widget = self.master.nametowidget(self._source + "_" + row)
    elif (self._source in self.master.children):
      source_widget = self.master.nametowidget(self._source)
    if source_widget:
      # special case - show all lists in a sharepoint site
      if self._source == 'sharepoint':
        from sp_custom import sp_lists
        self.setValues(sp_lists(source_widget.username, source_widget.password))
      else:
        self.setValues(smartfilelist.get(source_widget.get(), self._alternate))
    else:
      self.setValues([self._source])

    # reset the cursor back to default
    self._control['cursor'] = ''

class FileEntry(ttk.Frame):
  '''custom Entry, with label and a Browse button'''
  _label = None
  _button = None
  _control = None
  def __init__(self, master, label, wildcard=''):
    ttk.Frame.__init__(self, master, name=label)
    self._button = ttk.Button(self, text="⛘", command=self.onBrowse)
    self._button.pack(side=tk.RIGHT)
    self._output = False
    ttk.Style().configure('red.TButton', foreground='red')
    if isinstance(master, tkTable):
      self._control = ttk.Combobox(self)
    else:
      self._control = ttk.Combobox(self, width=-60)
      self._label = ttk.Label(self, text=label, width=-20)
      self._label.pack(side=tk.LEFT)
      self._output = self._label['text'].startswith("output")
    self._control.pack(expand=True, fill=tk.BOTH, side=tk.RIGHT)
    self._control.bind("<ButtonPress>", self.onButtonPress)
    self._wildcard_list = []
    self._wildcard_full = [("*", "*")]
    if len(wildcard):
      self._wildcard_list.extend(wildcard.split(','))
      self._wildcard_full.insert(0, (wildcard, ['*.' + _ for _ in self._wildcard_list]))
  
  # activate the browse button, which shows a native fileopen dialog and sets the Entry control
  def onBrowse(self):
    if self._output:
      flist = (filedialog.asksaveasfilename(filetypes=self._wildcard_full),)
    else:
      flist = filedialog.askopenfilenames(filetypes=self._wildcard_full)
    
    if(isinstance(flist, tuple)):
      if isinstance(self.master, tkTable) and len(flist) > 1:
        # instead of setting only the current control, call our parent to set multiple at once
        self.master.set(map(relative_paths, flist))
      else:
        self.set(",".join(map(relative_paths, flist)))

  def onButtonPress(self, *args):
    # temporarily set the cursor to a hourglass
    if (len(self._control['values'])):
      return

    self._control['cursor'] = 'watch'
    if len(self._wildcard_list):
      self._control['values'] = [_ for _ in os.listdir('.') if re.search('\.(?:' + '|'.join(self._wildcard_list) + ')$', _)]
    else:
      self._control['values'] = os.listdir('.')

    # reset the cursor back to default
    self._control['cursor'] = ''
    # reset red button back to black
    self._button['style'] = ''

  def get(self):
    return self._control.get()
  
  def set(self, value):
    if(value == None or str(value) == 'nan' or len(value) == 0):
      return
    if not self._output:
      self._button['style'] = '' if os.path.exists(value) else 'red.TButton'

    self._control.delete(0, tk.END)
    self._control.insert(0, value)

  def configure(self, **kw):
    if self._label is not None:
      self._label.configure(**kw)
    self._button.configure(**kw)
    self._control.configure(**kw)

# helper function to securely handle username + password
class Credentials(list):
  '''
  store credentials in a manageable manner
  '''
  _delim = ':'
  _encoding = 'utf-8'
  def __init__(self, s = None):
    super().__init__(('', ''))
    import base64
    import cryptography.fernet
    import uuid
    # use symmetric encryption on the password
    # local unique node id is used as key
    # decoding is possible to anyone with filesystem
    # and execute access to where the store was created!
    key = uuid.getnode().to_bytes(32, 'big')
    self._f = cryptography.fernet.Fernet(base64.urlsafe_b64encode(key))
    if s is not None:
      self.parse(s)

  def __str__(self):
    return self._delim.join((self.username, self.passhash))

  def parse(self, s):
    data = s.split(self._delim)
    if len(data) > 0:
      self.username = data[0]
    if len(data) > 1:
      self.passhash = data[1]

  @property
  def username(self):
    return self[0]

  @username.setter
  def username(self, value):
    self[0] = value

  @property
  def password(self):
    return self[1]

  @password.setter
  def password(self, value):
    self[1] = value

  @property
  def passhash(self):
    step1 = bytes(self[1], self._encoding)
    step2 = self._f.encrypt(step1)
    return str(step2, self._encoding)

  @passhash.setter
  def passhash(self, value):
    step1 = bytes(value, self._encoding)
    try:
      step2 = self._f.decrypt(step1)
    except:
      step2 = b''
    self[1] = str(step2, self._encoding)

# username + password
class CredentialsInput(ttk.LabelFrame):
  ''' 
  Control to input username and Password
  The result will be a single string in the format:
  user:encodedpass
  '''
  _delim = ':'
  def __init__(self, master, label, uri_value = None, pid_value = None):
    # create a container frame for the combo and label
    ttk.Labelframe.__init__(self, master, name=label, text=label)
    self.columnconfigure(1, weight=1)

    ttk.Label(self, text='👤').grid(row=0, column=0)
    ttk.Label(self, text='🔑').grid(row=1, column=0)
    
    self._u = ttk.Entry(self)
    self._u.grid(sticky=tk.W + tk.E, padx=4, pady=2, row=0, column=1)
    self._p = ttk.Entry(self, show='*')
    self._p.grid(sticky=tk.W + tk.E, padx=4, pady=2, row=1, column=1)
    self._c = Credentials()

  @property
  def username(self):
    return self._u.get()
    
  @property
  def password(self):
    return self._p.get()
    
  def get(self):
    self._c.username = self._u.get()
    self._c.password = self._p.get()
    return str(self._c)

  def set(self, value):
    if(value == None or len(value) == 0):
      return
    self._c.parse(value)
    if len(value) > 0:
      self._u.delete(0, tk.END)
      self._u.insert(0, self._c.username)
    if len(value) > 1:
      self._p.delete(0, tk.END)
      self._p.insert(0, self._c.password)

  def configure(self, **kw):
    self._u.configure(**kw)
    self._p.configure(**kw)

# template for entry + callback button to be overriden by subclass
class ButtonEntry(ttk.Frame):
  ''' should behave the same as Tix LabelEntry but with some customizations '''
  _label = None
  _control = None
  def __init__(self, master, label, callback = None):
    # create a container frame for the combo and label
    ttk.Frame.__init__(self, master, name=label)
    self._callback = callback
    self._control = ttk.Entry(self)
    self._control.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)
    if len(label) == 1:
      self._button = ttk.Button(self, text=label, command=self.action, width=3)
      self._button['style'] = 'basic.TButton'
    else:
      self._button = ttk.Button(self, text=label, command=self.action, width=16)

    self._button.pack(side=tk.RIGHT)
    
  def get(self):
    return self._control.get()
   
  def set(self, value):
    if value is None:
      return
    if not isinstance(value, str):
      value = str(value)

    self._control.delete(0, tk.END)
    self._control.insert(0, value)

  def action(self):
    if callable(self._callback):
      self.set(self._callback())

# create a table of entry/combobox widgets
class tkTable(ttk.Labelframe):
  def __init__(self, master, label, columns):
    ttk.Labelframe.__init__(self, master, name=label, text=label)

    self._label = label
    if len(self._label) == 0:
      self._label = str(self.winfo_id())
    self._columns = [UsageToken(_) for _ in columns]
    self._cells = []
    for i in range(len(self._columns)):
      self.columnconfigure(i, weight=1)
      ttk.Label(self, text=self._columns[i].name).grid(row=0, column=i)

    self.addRow()
    
  # return the table data as a serialized commalist
  def get(self, row=None, col=None):
    value = ""
    # retrieve all values as a 2d list
    if(row==None and col==None):
      value = commalist()
      for i in range(len(self._cells)):
        value.append([self.get(i, j) for j in range(len(self._columns))])
      # trim empty rows at the end
      for i in range(len(value)-1,0,-1):
        for j in range(len(value[i])):
          if value[i][j]:
            # outter loop will also break
            break
        else:
          value.pop()
          continue
        break
  
    elif(row < len(self._cells) and col < len(self._columns)):
      value = self._cells[row][col].get()
    return value

  # set the widget values, expanding the table rows as needed
  # input data must be a string containing a serialized commalist
  def set(self, data, row=None, col=0):
    if row is None:
      data = commalist().parse(data)
      for i in range(len(data)):
        if(isinstance(data[i], list)):
          for j in range(len(data[i])):
            self.set(data[i][j], i, j)
        else:
          self.set(data[i], i)
    else:
      # expand internal array to fit the data
      for i in range(len(self._cells), row+1):
        self.addRow()
      self._cells[row][col].set(data)

  def addRow(self):
    row = len(self._cells)
    self._cells.append([])
    for col in range(len(self._columns)+1):
      child = None
      if col == len(self._columns):
        if row == 0:
          child = ttk.Button(self, text="➕", width=3, command=self.addRow)
        else:
          child = ttk.Button(self, text="✖", width=3, command=lambda: self.delRow(row))
      else:
        token = self._columns[col]
        if(token.type == '@'):
          child = CheckBox(self, "%s_%s" % (token.name,row))
        elif(token.type == '*'):
          child = FileEntry(self, "%s_%s" % (token.name,row), token.data)
        elif(token.type == '='):
          child = LabelCombo(self, "%s_%s" % (token.name,row), token.data)
        elif(token.type == ':'):
          child = ComboPicker(self, "%s_%s" % (token.name,row), token.data)
        elif(token.type == '!'):
          child = ComboPicker(self, "%s_%s" % (token.name,row), token.data, True)
        else:
          child = LabelEntry(self, "%s_%s" % (token.name,row))
      child.grid(row=row+1, column=col, sticky="we")
      self._cells[row].append(child)
  
  def delRow(self, index=0):
    rows = self.get()
    if index < len(rows):
      del rows[index]
    self.clear()
    self.set(rows)
  
  def clear(self):
    for i in range(len(self._cells)-1,-1,-1):
      for j in range(len(self._cells[i])-1,-1,-1):
        self._cells[i][j].destroy()
    del self._cells[:]

  def configure(self, **kw):
    if "state" in kw:
      for v in self.children.values():
        v.configure(**kw)
    else:
      super().configure(**kw)

# main frame
class AppTk(tk.Tk):
  '''TK-Based Data driven GUI application'''
  _iconfile = None
  _logofile = None
  def __init__(self, usage, client=sys.argv[0]):
    ClientScript.init(client)
    tk.Tk.__init__(self)
    self.title(ClientScript._base)
    
    self._iconfile = Branding().name
    self._logofile = Branding('png', (100,100))
    self.iconbitmap(default=self._iconfile)

    self.columnconfigure(0, weight=1)
    self.canvas = tk.Canvas(width=self.winfo_screenwidth() * 0.35)
    self.script = ScriptFrame(self.canvas, usage)

    self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    self.canvas_frame = self.canvas.create_window((0,0), window=self.script, anchor="nw")
    self.vsb = ttk.Scrollbar(orient=tk.VERTICAL, command=self.canvas.yview)
    self.canvas.configure(yscrollcommand=self.vsb.set)

    self.vsb.pack(side=tk.LEFT, fill=tk.Y)
    self.script.bind("<Configure>", self.onFrameConfigure)
    self.canvas.bind('<Configure>', self.onCanvasConfigure)

    ttk.Label(self, text=ClientScript.header()).pack(side=tk.BOTTOM)
    
    self.logo = tk.Canvas(self, width=self._logofile.image.size[0], height=self._logofile.image.size[1])
    
    self.logo.create_image(0, 0, anchor='nw', image=self._logofile.photoimage)
    self.logo.pack(side=tk.TOP, anchor="ne")

    self.button = ttk.Button(self, text="Run", command=self.runScript)
    self.button.pack(side=tk.LEFT)
    
    self.progress = ttk.Progressbar(self, mode="determinate")
    self.progress.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=10)

    self.createMenu()
    self.script.set(Settings().load())

  def onCanvasConfigure(self, event):
    self.canvas.itemconfig(self.canvas_frame, width = event.width - 4)

  def onFrameConfigure(self, event):
    '''Reset the scroll region to encompass the inner frame'''
    self.canvas.configure(scrollregion=self.canvas.bbox("all"))
    self.canvas['height'] = min(self.winfo_screenheight() * 0.8, self.script.winfo_reqheight())

  def createMenu(self):
    '''create a hardcoded menu for our app'''
    # disable a legacy option to detach menus
    self.option_add('*tearOff', False)
    menubar = tk.Menu(self)
    menu_file = tk.Menu(menubar)
    menu_edit = tk.Menu(menubar)
    menu_help = tk.Menu(menubar)
    menubar.add_cascade(menu=menu_file, label='File')
    menubar.add_cascade(menu=menu_help, label='Help')
    menu_file.add_command(label='Copy Command Line', command=self.script.copy)
    menu_file.add_command(label='Open Settings', command=self.openSettings)
    menu_file.add_command(label='Save Settings', command=self.saveSettings)
    menu_file.add_command(label='Exit', command=self.destroy)
    menu_help.add_command(label='Help', command=self.showHelp)
    menu_help.add_command(label='Start Command Line Window', command=self.openCmd)
    menu_help.add_command(label='About', command=self.showAbout)
    self['menu'] = menubar
      
  def runScript(self):
    # run the process in another thread as not to block the GUI message loop
    def fork():
      self.button.configure(state = "disabled")
      self.progress.configure(value = 0, mode = "indeterminate")
      self.progress.start()
      
      if ClientScript.run(self.script):
        messagebox.showwarning(message="Check console messages",title=ClientScript.type())

      self.progress.stop()
      self.progress.configure(value = 100)
      self.progress.configure(mode = "determinate")
      self.button.configure(state = "enabled")

    threading.Thread(None, fork).start()

  def showHelp(self):
    script_pdf = ClientScript.file('pdf')
    if os.path.exists(script_pdf):
      os.system(script_pdf)
    else:
      messagebox.showerror('Help', 'Documentation file not found')
  
  def openCmd(self):
    print("starting cmd")
    os.system('start cmd /k python -V')
    
  def showAbout(self):
    messagebox.showinfo('About', 'Graphic User Interface to command line scripts')

  def openSettings(self):
    result = filedialog.askopenfilename(filetypes=[("ini", "*.ini")])
    if len(result) == 0:
      return
    self.script.set(Settings(result).load())
    
  def saveSettings(self):
    result = filedialog.asksaveasfilename(filetypes=[("ini", "*.ini")])
    if len(result) == 0:
      return
    Settings(result).save(self.script.get(True))
  
  def destroy(self):
    Settings().save(self.script.get(True))
    os.remove(self._iconfile)
    tk.Tk.destroy(self)

# images for window icon and watermark
class Branding(object):
  _gc = []
  def __init__(self, f='ICO', size=None, choice=None):
    if choice is None:
      self._choice = os.environ.get('USERDOMAIN')
    else:
      self._choice = choice

    self._format = f
    
    self._image = Image.new('RGBA', (800, 800))
    draw = ImageDraw.Draw(self._image)

    if self._choice=='VALENET':
      # Vale logo
      draw.polygon((430, 249, 397,698, 803,160,  766,139, 745,132, 727,130, 692,130, 655,142, 618,160, 571,188, 524,216, 477,236), "#eaab13")
      draw.chord((-80,105, 413,588), 228, 312, "#008f83")
      draw.polygon((0,165, 397,698, 720,270, 454,270, 454,270, 429,248, 328,165), "#008f83")
      draw.chord((403,-40, 770,327), 44, 136, "#eaab13")
    else:
      # open souce logo
      draw.pieslice([40, 40, 760, 760], 110, 70, '#3fa648')
      draw.ellipse([288, 288, 512, 512], Image.ANTIALIAS)

    del draw
    if size:
      self._image = self._image.resize(size)

  @property
  def file(self):
    import tempfile
    self._file = tempfile.NamedTemporaryFile(delete=False)
    self._image.save(self._file, self._format)
    return self._file

  @property
  def name(self):
    return self.file.name

  @property
  def image(self):
    return(self._image)
  
  @property
  def format(self):
    return(self._format)

  @property
  def data(self):
    import base64
    from io import BytesIO
    buffered = BytesIO()
    self._image.save(buffered, format=self._format)
    return base64.b64encode(buffered.getvalue())

  @property
  def photoimage(self):
    # if we dont store the image in a property, it will be garbage colected before being displayed
    self._pi = tk.PhotoImage(data=self.data)
    return self._pi

### } GUI

### { ENTRY POINTS

# default main for when this script is standalone
# when this as a library, will redirect to the caller script main()
def main(*args):
  if (__name__ != '__main__'):
    from __main__ import main
    # redirect to caller script main
    main(*args)
    return

# special entry point for cmd
if __name__ == '__main__' and sys.argv[0].endswith('_gui.py') and len(sys.argv) == 1:
  pass
elif __name__ == '__main__' and len(sys.argv) == 2:
  AppTk(None, sys.argv[1]).mainloop()
elif __name__ == '__main__' and sys.argv[0].endswith('_gui.py'):
  main(sys.argv)

### } ENTRY POINTS
